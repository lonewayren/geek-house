#!/usr/bin/env python
# -*- coding:utf-8 -*- 
# __author__ = 'renxiang03'
from __future__ import unicode_literals
import os
import traceback
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from apps.movie.serializers import Movie, Link, Source, Category, MovieListSerializer, LinkSerializer, CategorySerializer
from django.db import transaction
from django.conf import settings

ERROR_LIST = []


def load_data():
    filename = os.path.join(settings.BASE_DIR, 'apps/spider/movie.xlsx')
    wb = load_workbook(filename, data_only=True)
    for ws in wb.worksheets:
        source = Source.objects.first()
        line = 1
        num = True
        while num:
            line += 1
            num = ws['A' + str(line)].value
            try:
                num = ws['A'+str(line)].value
                if not num:
                    break
                if not num.isdigit():
                    continue
                title = ws['B'+str(line)].value.strip() if ws['B'+str(line)].value else ''
                language = ws['D'+str(line)].value.strip() if ws['D'+str(line)].value else ''
                category = ws['E'+str(line)].value.strip() if ws['E'+str(line)].value else ''
                age = ws['F'+str(line)].value.strip() if ws['F'+str(line)].value else ''
                area = ws['G'+str(line)].value.strip() if ws['G'+str(line)].value else ''
                score = ws['H'+str(line)].value.strip() if ws['H'+str(line)].value else 0
                online_date = ws['O'+str(line)].value.strip() if ws['O'+str(line)].value else '1949-10-01'
                star = ws['Q'+str(line)].value.strip() if ws['Q'+str(line)].value else ''
                if '\n' in star:
                    star = ",".join([item.strip() for item in star.split('\n')[:3:]])
                description = ws['R'+str(line)].value.strip() if ws['R'+str(line)].value else ''
                cover = ws['S'+str(line)].value.strip() if ws['S'+str(line)].value else ''
                comments = ws['T'+str(line)].value.strip() if ws['T'+str(line)].value else ''
                if comments.startswith('//'):
                    comments = 'http:' + comments
                link_4k = ws['I'+str(line)].value.strip() if ws['I'+str(line)].value else ''
                link_hd = ws['J'+str(line)].value.strip() if ws['J'+str(line)].value else ''
                link_lg = ws['K'+str(line)].value.strip() if ws['K'+str(line)].value else ''
                link_sm = ws['L'+str(line)].value.strip() if ws['L'+str(line)].value else ''
                link_yb = ws['M'+str(line)].value.strip() if ws['M'+str(line)].value else ''
                link_dict = {
                    'sm': link_sm,
                    'lg': link_lg,
                    'hd': link_hd,
                    '4k': link_4k,
                    'yb': link_yb,
                }

                # print num, title, language, category, age, area, score, online_date, star, description, cover, comments
                with transaction.atomic():
                    if not Category.objects.filter(name=category).exists():
                        category_ins = CategorySerializer(data={
                            'name': category
                        }).shortcut_create()
                    else:
                        category_ins = Category.objects.get(name=category)
                    if Movie.objects.filter(source_uid=num).exists():
                        continue
                    movie_ins = MovieListSerializer(data={
                        'title': title,
                        'star': star,
                        'cover': cover,
                        'language': language,
                        'age': age,
                        'area': area,
                        'score': score,
                        'online_date': online_date,
                        'comments': comments,
                        'category': [category_ins.id],
                        'description': description,
                        'source_uid': num,
                        'source': source.id,
                    }).shortcut_create()
                    for dpi, link in link_dict.items():
                        if link and not Link.objects.filter(movie=movie_ins.id, definition=dpi).exists():
                            link_ins = LinkSerializer(data={
                                'movie': movie_ins.id,
                                'link': link,
                                'type': 'xl',
                                'definition': dpi,
                                'secret': '',
                                'valid': True,
                            }).shortcut_create()
            except Exception as e:
                print num, online_date, comments, category, online_date, star
                traceback.print_exc()
                ERROR_LIST.append(num)
            # break
    print '*'*20+'导入错误'+'*'*20
    for num in ERROR_LIST:
        print num


class Command(BaseCommand):
    def handle(self, *args, **options):
        load_data()
