#!/usr/bin/env python
# -*- coding:utf-8 -*-
from __future__ import unicode_literals
"""
Created on Thu Feb  8 11:44:17 2018

@author: 周立凤
"""
from openpyxl import Workbook, load_workbook


class HtmlOutputer(object):
    def __init__(self):
        self.success_list = []
        self.failure_list = []
        self.filename = './采集结果.xlsx'
        self.id_list = []
        self.not_found_list = []

    @property
    def sheetname(self):
        return '%s-%s' % (min(self.id_list), max(self.id_list))

    def collect_data(self, book):
        self.id_list.append(book.source_id)
        if book.success:
            self.success_list.append(book)
        else:
            self.failure_list.append(book)

    def output(self):
        wb = load_workbook(self.filename)
        ws = wb.create_sheet(self.sheetname)

        first_row = ['ID','书名','分类','简介','提取码','百度网盘链接','封面']
        ws.append(first_row)
        for book in self.success_list:
            ws.append([getattr(book, attr) for attr in book.attr_list])
        wb.save(self.filename)

    def add_not_found(self, url):
        self.not_found_list.append(url)

    def statis(self):
        print '*' * 20, '失败列表', '*' * 20
        for item in self.failure_list:
            print item.source_url
        print '*' * 20, '统计信息', '*' * 20
        print "合计:%s" % (self.success_list.__len__() + self.failure_list.__len__() + self.not_found_list.__len__())
        print "成功:%s" % self.success_list.__len__()
        print "失败:%s" % self.failure_list.__len__()
        print "404:%s" % self.not_found_list.__len__()

