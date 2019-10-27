#!/usr/bin/env python
# -*- coding:utf-8 -*- 
# __author__ = 'renxiang03'
from __future__ import unicode_literals
import os
import urlparse
from openpyxl import load_workbook
from django.core.management.base import BaseCommand, CommandError
from apps.eBook.serializers import Book, Link, Source, Category, BookListSerializer, LinkSerializer, CategorySerializer
from django.db import transaction
from django.conf import settings

ERROR_LIST = []

def format_url(raw_url):
    raw_url = urlparse.unquote(raw_url)
    if 'surl=' in raw_url:
        qs = urlparse.parse_qs(urlparse.urlparse(raw_url).query)
        surl = qs.get('surl', ['xxx'])[0]
        raw_url = 'https://pan.baidu.com/s/1%s' % surl
    raw_url = raw_url.strip()
    if '#' in raw_url:
        raw_url = raw_url.split('#')[0]
    return raw_url


def load_data():
    filename = os.path.join(settings.BASE_DIR, 'apps/spider/采集结果.xlsx')
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        print ws.title
        if ws.title.endswith('$$$$'):
            continue
        source = Source.objects.first()
        line = 0
        num = True
        while num:
            line += 1
            num = ws['A' + str(line)].value
            # print line,
            try:
                num = ws['A'+str(line)].value
                if not num:
                    break
                if not num.isnumeric():
                    continue
                title = ws['B'+str(line)].value.strip() if ws['B'+str(line)].value else ''
                category = ws['C'+str(line)].value.strip() if ws['C'+str(line)].value else ''
                description = ws['D'+str(line)].value.strip() if ws['D'+str(line)].value else ''
                secret = ws['E'+str(line)].value.strip() if ws['E'+str(line)].value else ''
                link = ws['F'+str(line)].value.strip() if ws['F'+str(line)].value else ''
                cover = ws['G'+str(line)].value.strip() if ws['G'+str(line)].value else ''
                # print num, title, category, secret, link, cover

                with transaction.atomic():
                    if Book.objects.filter(source_uid=num).exists():
                        clean_link = format_url(link)
                        book = Book.objects.filter(source_uid=num).first()
                        old_links = Link.objects.filter(book=book.id, type='bd', secret=secret).all()
                        if not old_links.exists():
                            # print
                            # print 'no link, %s, %s' % (num, title)
                            continue
                        old_link = old_links.first()
                        if old_link.link.split('://')[-1] != clean_link.split('://')[-1]:
                            print 'modify link, linkID({id}) bookID({bid} old_link({old_link}) new_link({new_link})'.format(id=old_link.id, bid=book.id, old_link=old_link.link, new_link=clean_link)
                            old_link.link = clean_link
                            old_link.save()
                        continue
                    if not Category.objects.filter(name=category).exists():
                        category_ins = CategorySerializer(data={
                            'name': category
                        }).shortcut_create()
                    else:
                        category_ins = Category.objects.get(name=category)
                    book_ins = BookListSerializer(data={
                        'title': title,
                        'cover': cover,
                        'category': [category_ins.id],
                        'description': description,
                        'source_uid': num,
                        'source': source.id,
                    }).shortcut_create()
                    link_ins = LinkSerializer(data={
                        'book': book_ins.id,
                        'link': link,
                        'type': 'bd',
                        'secret': secret,
                    }).shortcut_create()
            except Exception as e:
                import traceback;
                traceback.print_exc();
                ERROR_LIST.append(num)
        ws.title += '$'
    wb.save(filename)
    print '*'*20+'导入错误'+'*'*20
    for num in ERROR_LIST:
        print num


class Command(BaseCommand):
    def handle(self, *args, **options):
        load_data()
